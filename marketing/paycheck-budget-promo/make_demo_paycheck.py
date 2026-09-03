"""Fill the Paycheck Budget template with one demo month (July 2026).

Usage: python3 make_demo_paycheck.py \"Paycheck Budget  Light.xlsx\" PaycheckBudget_Light_DEMO.xlsx
"""
import datetime as dt
import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties

import sys
SRC = sys.argv[1]                       # the clean template
OUT = sys.argv[2]                       # workbook with the demo month in it

wb = openpyxl.load_workbook(SRC)

# ---------- Setup: a couple more funds / debts ----------
su = wb["Setup"]
su["H10"] = "Vacation"          # savings funds
su["H11"] = "Home Repairs"
su["J9"]  = "Student loan"      # debts

# Dashboard A4 holds the label "Month", but column A is the 2-char gutter and
# B4 carries the value, so the label renders clipped to "M". Cleared here so it
# stays off camera — in the template itself the label wants a wider home.
wb["Dashboard"]["A4"] = None

# ---------- Paycheck Budget: planned split P1 / P2 ----------
pb = wb["Paycheck Budget"]
planned = [
    ("Rent or Mortgage", 1450, 0),
    ("Utilities",          90, 90),
    ("Groceries",         260, 260),
    ("Transport",         120, 120),
    ("Dining out",         80, 70),
    ("Subscriptions",      45, 0),
    ("Health",             60, 40),
    ("Insurance",         145, 0),
    ("Phone and Internet",  0, 85),
    ("Shopping",          100, 80),
    ("Fun and Leisure",    90, 70),
    ("Gifts",              40, 30),
]
for i, (_, p1, p2) in enumerate(planned):
    r = 12 + i
    pb.cell(row=r, column=2, value=p1)
    pb.cell(row=r, column=3, value=p2)

# ---------- Transactions ----------
tx = wb["Transactions"]
D = lambda d: dt.datetime(2026, 7, d)
rows = [
    (D(3),  "Income",   "Paycheck 1",        "Checking", "1",     2100.00, "Salary"),
    (D(3),  "Expense",  "Rent or Mortgage",  "Checking", "1",     1450.00, "July rent"),
    (D(3),  "Transfer", "Emergency Fund",    "Savings",  "1",      250.00, "Pay yourself first"),
    (D(3),  "Transfer", "Vacation",          "Savings",  "1",      120.00, "Greece 2027"),
    (D(4),  "Expense",  "Groceries",         "Checking", "1",      118.40, "Weekly shop"),
    (D(4),  "Expense",  "Dining out",        "Credit Card", "1",    28.50, "Brunch"),
    (D(5),  "Expense",  "Transport",         "Checking", "1",       62.00, "Fuel"),
    (D(6),  "Expense",  "Subscriptions",     "Credit Card", "1",    15.99, "Streaming"),
    (D(6),  "Expense",  "Phone and Internet","Checking", "1",       45.00, "Mobile"),
    (D(7),  "Expense",  "Utilities",         "Checking", "1",       88.20, "Electricity"),
    (D(8),  "Expense",  "Health",            "Checking", "1",       34.00, "Pharmacy"),
    (D(9),  "Expense",  "Groceries",         "Checking", "1",      104.75, "Weekly shop"),
    (D(10), "Expense",  "Fun and Leisure",   "Cash",     "1",       42.00, "Cinema"),
    (D(11), "Expense",  "Insurance",         "Checking", "1",      145.00, "Car insurance"),
    (D(11), "Expense",  "Shopping",          "Credit Card", "1",    76.30, "Summer shoes"),
    (D(12), "Expense",  "Transport",         "Checking", "1",       58.40, "Fuel"),
    (D(13), "Expense",  "Dining out",        "Credit Card", "1",    31.80, "Pizza night"),
    (D(14), "Expense",  "Groceries",         "Checking", "1",       96.10, "Weekly shop"),
    (D(15), "Expense",  "Gifts",             "Cash",     "1",       35.00, "Birthday gift"),
    (D(16), "Expense",  "Utilities",         "Checking", "1",       26.50, "Water"),
    (D(17), "Income",   "Paycheck 2",        "Checking", "2",     2100.00, "Salary"),
    (D(17), "Transfer", "Emergency Fund",    "Savings",  "2",      200.00, "Pay yourself first"),
    (D(17), "Transfer", "New Car",           "Savings",  "2",      300.00, "Down payment"),
    (D(17), "Transfer", "Holiday",           "Savings",  "2",      150.00, "December fund"),
    (D(18), "Expense",  "Groceries",         "Checking", "2",      112.90, "Weekly shop"),
    (D(18), "Expense",  "Phone and Internet","Checking", "2",       40.00, "Internet"),
    (D(19), "Expense",  "Transport",         "Checking", "2",       55.00, "Fuel"),
    (D(19), "Expense",  "Fun and Leisure",   "Credit Card", "2",    48.00, "Concert tickets"),
    (D(20), "Expense",  "Subscriptions",     "Credit Card", "2",    12.99, "Music"),
    (D(21), "Expense",  "Health",            "Checking", "2",       60.00, "Dentist"),
    (D(22), "Income",   "Extra Income",      "Checking", "Extra",   320.00, "Freelance invoice"),
    (D(22), "Transfer", "Vacation",          "Savings",  "Extra",   150.00, "Flights"),
    (D(22), "Expense",  "Groceries",         "Checking", "2",       88.60, "Weekly shop"),
    (D(23), "Expense",  "Dining out",        "Credit Card", "2",    26.40, "Lunch out"),
    (D(24), "Expense",  "Shopping",          "Credit Card", "2",    54.20, "Household"),
    (D(25), "Expense",  "Utilities",         "Checking", "2",       61.30, "Gas"),
    (D(26), "Expense",  "Transport",         "Checking", "2",       32.00, "Train pass"),
    (D(27), "Expense",  "Groceries",         "Checking", "2",       79.45, "Weekly shop"),
    (D(27), "Expense",  "Gifts",             "Cash",     "2",       25.00, "Wedding card"),
    (D(28), "Expense",  "Fun and Leisure",   "Cash",     "2",       36.00, "Museum"),
    (D(29), "Expense",  "Dining out",        "Credit Card", "2",    18.90, "Coffee run"),
    (D(30), "Transfer", "Home Repairs",      "Savings",  "2",       60.00, "Boiler service"),
]
for i, r in enumerate(rows):
    for j, v in enumerate(r):
        tx.cell(row=5 + i, column=1 + j, value=v)

# ---------- Bill Calendar ----------
bc = wb["Bill Calendar"]
bills = [
    ("Rent or Mortgage",  1450.00, "Monthly",  1,  "Yes"),
    ("Electricity",         88.20, "Monthly",  7,  "Yes"),
    ("Water",               26.50, "Monthly", 16,  "Yes"),
    ("Gas",                 61.30, "Monthly", 25,  "Yes"),
    ("Mobile plan",         45.00, "Monthly",  6,  "Yes"),
    ("Internet",            40.00, "Monthly", 18,  "Yes"),
    ("Car insurance",      145.00, "Monthly", 11,  "Yes"),
    ("Streaming",           15.99, "Monthly",  6,  "Yes"),
    ("Music",               12.99, "Monthly", 20,  "Yes"),
    ("Gym",                 24.00, "Weekly",   5,  "No"),
    ("Home insurance",     420.00, "Yearly",  14,  "No"),
    ("Cloud storage",        9.99, "Monthly", 22,  "No"),
]
for i, (name, amt, freq, day, paid) in enumerate(bills):
    r = 5 + i
    bc.cell(row=r, column=1, value=name)
    bc.cell(row=r, column=2, value=amt)
    bc.cell(row=r, column=3, value=freq)
    bc.cell(row=r, column=5, value=day)
    bc.cell(row=r, column=6, value=paid)

# ---------- Savings targets ----------
sv = wb["Savings"]
for i, target in enumerate([6000, 1200, 9000, 3500, 1500]):   # Emergency, Holiday, New Car, Vacation, Home Repairs
    sv.cell(row=5 + i, column=2, value=target)

# ---------- Debt ----------
db = wb["Debt"]
debts = [
    (2450.00, 1180.00, 95.00),    # Credit Card
    (9800.00, 3600.00, 245.00),   # Car loan
    (14200.00, 5100.00, 190.00),  # Student loan
]
for i, (bal, paid, minp) in enumerate(debts):
    r = 5 + i
    db.cell(row=r, column=2, value=bal)
    db.cell(row=r, column=3, value=paid)
    db.cell(row=r, column=6, value=minp)

# ---------- page setup: one page per sheet ----------
areas = {
    "Start Here":      "A1:H24",
    "Setup":           "A1:K26",
    "Dashboard":       "A1:O34",
    "Paycheck Budget": "A1:G33",
    "Transactions":    "A1:G34",
    "Bill Calendar":   "A1:F18",
    "Savings":         "A1:E12",
    "Debt":            "A1:F13",
}
for ws in wb.worksheets:
    ws.print_area = areas[ws.title]
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = ws.page_margins.right = 0.15
    ws.page_margins.top = ws.page_margins.bottom = 0.15
    ws.page_margins.header = ws.page_margins.footer = 0
    ws.print_options.horizontalCentered = False
    ws.sheet_view.showGridLines = False
    ws.oddHeader.left.text = ws.oddHeader.center.text = ws.oddHeader.right.text = ""
    ws.oddFooter.left.text = ws.oddFooter.center.text = ws.oddFooter.right.text = ""

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved", OUT)
