"""Render every sheet of the demo workbook to a cropped PNG, plus the
Weekly->Monthly variant of Bill Calendar and the cell geometry the video needs.

Usage: WORK=./work python3 render_sheets.py PaycheckBudget_Light_DEMO.xlsx
Needs libreoffice-calc (headless) with recalculation on load enabled.
"""
import json, os, shutil, subprocess, sys
import numpy as np
import openpyxl, pymupdf
from PIL import Image

D  = os.environ.get("WORK", os.path.dirname(os.path.abspath(__file__))+"/work")
HOME = D + "/lohome"
SRCWB = sys.argv[1] if len(sys.argv) > 1 else D+"/wb/demo.xlsx"
os.makedirs(D+"/wb", exist_ok=True)
os.makedirs(HOME, exist_ok=True)
NAMES = ["Start Here","Setup","Dashboard","Paycheck Budget","Transactions","Bill Calendar","Savings","Debt"]
DPI = 200

# ---- variant B: Gym bill switched Weekly -> Monthly ----
shutil.copy(SRCWB, D+"/wb/demo.xlsx")
wb = openpyxl.load_workbook(D+"/wb/demo.xlsx")
bc = wb["Bill Calendar"]
for r in range(5, 17):
    if bc.cell(row=r, column=1).value == "Gym":
        bc.cell(row=r, column=3, value="Monthly")
wb.save(D+"/wb/demo_b.xlsx")

def ensure_profile():
    """LibreOffice keeps Excel's cached values unless told to recalculate."""
    prof = HOME + "/loprofile"
    reg = prof + "/user/registrymodifications.xcu"
    if not os.path.exists(reg):
        subprocess.run(["soffice", "--headless", "--terminate_after_init",
                        "-env:UserInstallation=file://" + prof],
                       env=dict(os.environ, HOME=HOME), capture_output=True, timeout=180)
    txt = open(reg, encoding="utf-8").read()
    if "OOXMLRecalcMode" not in txt:
        item = ('<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
                '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n')
        open(reg, "w", encoding="utf-8").write(txt.replace("</oor:items>", item + "</oor:items>"))

ensure_profile()
env = dict(os.environ, HOME=HOME)
shutil.rmtree(D+"/wb/pdf", ignore_errors=True)
for f in ("demo.xlsx", "demo_b.xlsx"):
    subprocess.run(["soffice","--headless","--norestore",
                    "-env:UserInstallation=file://"+HOME+"/loprofile",
                    "--convert-to","pdf","--outdir",D+"/wb/pdf",D+"/wb/"+f],
                   env=env, check=True, capture_output=True, timeout=300)

os.makedirs(D+"/screens", exist_ok=True)
meta = {"dpi": DPI, "sheets": {}}
docA = pymupdf.open(D+"/wb/pdf/demo.pdf")
docB = pymupdf.open(D+"/wb/pdf/demo_b.pdf")
s = DPI/72.0

for i, name in enumerate(NAMES):
    pixA = docA[i].get_pixmap(dpi=DPI)
    imA = Image.frombytes("RGB", (pixA.width, pixA.height), pixA.samples)
    a = np.asarray(imA).astype(int)
    ys, xs = np.where(a.sum(axis=2) < 750)
    pad = 24
    x0, x1 = max(int(xs.min())-pad, 0), min(int(xs.max())+pad, imA.width)
    y0, y1 = max(int(ys.min())-pad, 0), min(int(ys.max())+pad, imA.height)
    key = f"{i}_{name.replace(' ','_')}"
    imA.crop((x0,y0,x1,y1)).save(f"{D}/screens/{key}.png")
    ent = {"crop":[x0,y0,x1,y1], "size":[x1-x0, y1-y0]}
    if name == "Bill Calendar":
        pixB = docB[i].get_pixmap(dpi=DPI)
        imB = Image.frombytes("RGB", (pixB.width, pixB.height), pixB.samples)
        imB.crop((x0,y0,x1,y1)).save(f"{D}/screens/{key}_b.png")
        # locate the clickable frequency cell ("Weekly" on the Gym row) in crop space
        rects = docA[i].search_for("Weekly")
        r = rects[0]
        ent["cell"] = [r.x0*s-x0, r.y0*s-y0, r.x1*s-x0, r.y1*s-y0]
        for label, q in (("cost_val", "$104.00"), ("total_val", "$2,033.96")):
            hit = docA[i].search_for(q)
            if hit:
                c = hit[0]
                ent[label] = [c.x0*s-x0, c.y0*s-y0, c.x1*s-x0, c.y1*s-y0]
    meta["sheets"][key] = ent
    print(key, ent["size"], ent.get("cell"))

json.dump(meta, open(D+"/screens/meta.json","w"), indent=1)
print("ok")
